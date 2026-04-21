"""
Export Excel optimisé — 2 requêtes DB au lieu de N×7
"""

from decimal import Decimal
from datetime import timedelta
from django.db.models import Sum, Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from ..agents.models import Agent, CreationMarchand, SuiviMarchand

JOURS_FR = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
OBJECTIF_OPENER = 5
SEUIL_OPENER = 3
TRANSPORT_BASE_OPENER = 6000
PLAFOND_ANIMATEUR = Decimal('50000')
TAUX_ANIMATEUR = Decimal('0.10')

# ──────────────────────────────────────────────
# Styles partagés
# ──────────────────────────────────────────────
HDR_FILL = PatternFill(start_color="0A2D6E", end_color="0A2D6E", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True)
ACC_FILL = PatternFill(start_color="FFB800", end_color="FFB800", fill_type="solid")
ACC_FONT = Font(color="0A2D6E", bold=True)
OK_FILL  = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
KO_FILL  = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
THIN     = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
NUM_FMT  = '#,##0'


def _f(v):
    """Convertit Decimal → float pour openpyxl."""
    return float(v) if isinstance(v, Decimal) else v


def _week_dates(date_debut, date_fin):
    d, dates = date_debut, []
    while d <= date_fin:
        dates.append(d)
        d += timedelta(days=1)
    return dates


def _hdr(ws, title, date_debut, date_fin):
    ws.merge_cells('A1:J1')
    ws['A1'] = f"MixxPay Agents — Rapport Transport {title}"
    ws['A1'].font = Font(size=14, bold=True, color="0A2D6E")
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Période : du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(size=11)
    ws.merge_cells('A3:J3')
    ws['A3'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A3'].font = Font(size=10, color="64748B")


def _auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                mx = max(mx, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(mx + 2, 35)


class ExcelExport:

    @staticmethod
    def export_openers(date_debut, date_fin):
        # ── 1 requête : toutes les créations de la semaine ──────────────
        rows = (
            CreationMarchand.objects
            .filter(date_activite__gte=date_debut, date_activite__lte=date_fin)
            .values('opener_id', 'date_activite')
            .annotate(nb=Count('id'))
        )
        # {opener_id: {date: nb}}
        detail = {}
        for r in rows:
            detail.setdefault(r['opener_id'], {})[r['date_activite']] = r['nb']

        # ── 1 requête : tous les openers ─────────────────────────────────
        openers = list(Agent.objects.filter(est_opener=True).order_by('numero'))

        dates = _week_dates(date_debut, date_fin)

        wb = Workbook()
        wb.remove(wb.active)
        ws_r = wb.create_sheet("Resume Openers")
        ws_d = wb.create_sheet("Detail Openers")

        _hdr(ws_r, "Openers", date_debut, date_fin)
        _hdr(ws_d, "Openers - Detail journalier", date_debut, date_fin)

        # ── Feuille Résumé ───────────────────────────────────────────────
        hdrs = ["Agent", "Equipe", "Team", "Objectif", "Seuil",
                "Realisation", "Taux (%)", "Transport base", "Transport net", "Statut"]
        row = 5
        for col, h in enumerate(hdrs, 1):
            c = ws_r.cell(row=row, column=col, value=h)
            c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
            c.alignment = Alignment(horizontal='center')

        total_transport = 0
        total_realisation = 0
        resume_rows = []

        for opener in openers:
            jour_data = detail.get(opener.id, {})
            realisation = sum(jour_data.values())
            transport = TRANSPORT_BASE_OPENER if realisation >= SEUIL_OPENER else 0
            taux = round(realisation / OBJECTIF_OPENER * 100, 1) if realisation else 0
            statut = "Atteint" if realisation >= SEUIL_OPENER else "Non atteint"
            total_transport += transport
            total_realisation += realisation
            resume_rows.append((opener, realisation, taux, transport, statut))

        for opener, realisation, taux, transport, statut in resume_rows:
            row += 1
            vals = [opener.numero, opener.equipe or '-', opener.team or '-',
                    OBJECTIF_OPENER, SEUIL_OPENER, realisation, taux,
                    TRANSPORT_BASE_OPENER, transport, statut]
            for col, v in enumerate(vals, 1):
                c = ws_r.cell(row=row, column=col, value=v)
                c.border = THIN
                if isinstance(v, (int, float)) and col not in (1, 2, 3, 10):
                    c.number_format = NUM_FMT
            ws_r.cell(row=row, column=10).fill = OK_FILL if statut == "Atteint" else KO_FILL

        row += 2
        ws_r.cell(row=row, column=1, value="TOTAUX").font = Font(bold=True)
        ws_r.cell(row=row, column=6, value=total_realisation)
        ws_r.cell(row=row, column=9, value=total_transport)
        ws_r.cell(row=row, column=9).number_format = NUM_FMT

        # ── Feuille Détail ───────────────────────────────────────────────
        cur = 5
        for opener, realisation, taux, transport, statut in resume_rows:
            jour_data = detail.get(opener.id, {})

            # Titre agent
            ws_d.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=4)
            c = ws_d.cell(row=cur, column=1,
                          value=f"Agent : {opener.numero}  |  Equipe : {opener.equipe or '-'}  |  Team : {opener.team or '-'}")
            c.font, c.fill = ACC_FONT, ACC_FILL
            cur += 1

            for col, h in enumerate(["Jour", "Date", "Creations"], 1):
                c = ws_d.cell(row=cur, column=col, value=h)
                c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
            cur += 1

            for d in dates:
                nb = jour_data.get(d, 0)
                ws_d.cell(row=cur, column=1, value=JOURS_FR[d.weekday()])
                ws_d.cell(row=cur, column=2, value=d.strftime('%d/%m/%Y'))
                ws_d.cell(row=cur, column=3, value=nb)
                cur += 1

            ws_d.cell(row=cur, column=1, value="TOTAL").font = Font(bold=True)
            ws_d.cell(row=cur, column=3, value=realisation)
            ws_d.cell(row=cur, column=4,
                      value=f"{transport:,} FCFA ({'Atteint' if realisation >= SEUIL_OPENER else 'Non atteint'})")
            cur += 2

        _auto_width(ws_r)
        _auto_width(ws_d)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="transport_openers_'
            f'{date_debut.strftime("%Y%m%d")}_{date_fin.strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response

    @staticmethod
    def export_animateurs(date_debut, date_fin):
        # ── 1 requête : volumes par animateur × jour ─────────────────────
        rows = (
            SuiviMarchand.objects
            .filter(date_activite__gte=date_debut, date_activite__lte=date_fin)
            .values('animateur_id', 'date_activite')
            .annotate(volume=Sum('montant'))
        )
        detail = {}
        for r in rows:
            detail.setdefault(r['animateur_id'], {})[r['date_activite']] = r['volume']

        # ── 1 requête : tous les animateurs ──────────────────────────────
        animateurs = list(Agent.objects.filter(est_animateur=True).order_by('numero'))

        dates = _week_dates(date_debut, date_fin)

        wb = Workbook()
        wb.remove(wb.active)
        ws_r = wb.create_sheet("Resume Animateurs")
        ws_d = wb.create_sheet("Detail Animateurs")

        _hdr(ws_r, "Animateurs", date_debut, date_fin)
        _hdr(ws_d, "Animateurs - Detail journalier", date_debut, date_fin)

        # ── Feuille Résumé ───────────────────────────────────────────────
        hdrs = ["Agent", "Volume realise (FCFA)", "Transport a payer (FCFA)"]
        row = 5
        for col, h in enumerate(hdrs, 1):
            c = ws_r.cell(row=row, column=col, value=h)
            c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
            c.alignment = Alignment(horizontal='center')

        total_vol = Decimal('0')
        total_tr  = Decimal('0')
        resume_rows = []

        for anim in animateurs:
            jour_data = detail.get(anim.id, {})
            vol_total = sum(jour_data.values(), Decimal('0'))
            transport = min(vol_total * TAUX_ANIMATEUR, PLAFOND_ANIMATEUR)
            total_vol += vol_total
            total_tr  += transport
            resume_rows.append((anim, vol_total, transport))

        for anim, vol_total, transport in resume_rows:
            row += 1
            vals = [anim.numero, _f(vol_total), _f(transport)]
            for col, v in enumerate(vals, 1):
                c = ws_r.cell(row=row, column=col, value=v)
                c.border = THIN
                if col > 1:
                    c.number_format = NUM_FMT

        row += 2
        ws_r.cell(row=row, column=1, value="TOTAUX").font = Font(bold=True)
        ws_r.cell(row=row, column=2, value=_f(total_vol)).number_format = NUM_FMT
        ws_r.cell(row=row, column=3, value=_f(total_tr)).number_format = NUM_FMT
        ws_r.cell(row=row, column=2).number_format = NUM_FMT
        ws_r.cell(row=row, column=3).number_format = NUM_FMT

        # ── Feuille Détail ───────────────────────────────────────────────
        cur = 5
        for anim, vol_total, transport in resume_rows:
            jour_data = detail.get(anim.id, {})

            ws_d.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=4)
            c = ws_d.cell(row=cur, column=1, value=f"Agent : {anim.numero}")
            c.font, c.fill = ACC_FONT, ACC_FILL
            cur += 1

            for col, h in enumerate(["Jour", "Date", "Volume (FCFA)", "Transport (FCFA)"], 1):
                c = ws_d.cell(row=cur, column=col, value=h)
                c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
            cur += 1

            for d in dates:
                v = jour_data.get(d, Decimal('0'))
                tr = min(v * TAUX_ANIMATEUR, PLAFOND_ANIMATEUR)
                ws_d.cell(row=cur, column=1, value=JOURS_FR[d.weekday()])
                ws_d.cell(row=cur, column=2, value=d.strftime('%d/%m/%Y'))
                ws_d.cell(row=cur, column=3, value=_f(v)).number_format  # set below
                ws_d.cell(row=cur, column=4, value=_f(tr))
                ws_d.cell(row=cur, column=3).number_format = NUM_FMT
                ws_d.cell(row=cur, column=4).number_format = NUM_FMT
                cur += 1

            ws_d.cell(row=cur, column=1, value="TOTAL").font = Font(bold=True)
            ws_d.cell(row=cur, column=3, value=_f(vol_total))
            ws_d.cell(row=cur, column=4, value=_f(transport))
            ws_d.cell(row=cur, column=3).number_format = NUM_FMT
            ws_d.cell(row=cur, column=4).number_format = NUM_FMT
            cur += 2

        _auto_width(ws_r)
        _auto_width(ws_d)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="transport_animateurs_'
            f'{date_debut.strftime("%Y%m%d")}_{date_fin.strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response
